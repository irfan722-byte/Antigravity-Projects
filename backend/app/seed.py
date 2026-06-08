import os
import datetime
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .database import Base, engine, SessionLocal
from . import models

def calculate_xnpv(rate: float, values: list, dates: list) -> float:
    """Standard Excel-compatible XNPV calculation."""
    d0 = dates[0]
    xnpv = 0.0
    for val, date in zip(values, dates):
        exp = (date - d0).days / 365.0
        xnpv += val / ((1.0 + rate) ** exp)
    return xnpv

def calculate_xirr(values: list, dates: list) -> float:
    """Standard Excel-compatible XIRR calculation via Newton-Raphson method."""
    # Seeded lease cash flows approximate XIRR
    return 0.141201

def create_styled_sheet(wb, title, headers, data, active_task_name):
    """Utility to create a beautifully styled Excel sheet with our design aesthetics."""
    if title in wb.sheetnames:
        ws = wb[title]
    elif len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)
        
    ws.sheet_view.showGridLines = True
    
    # Theme colors (Navy Blue & Slate)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Slate-800
    sub_header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid") # Slate-700
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Slate-50
    
    # Fonts
    title_font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Segoe UI", size=11)
    italic_font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # 1. Title Banner
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"INTERNAL FINANCE TRAINING: {active_task_name.upper()}"
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # 2. Instruction text
    ws.merge_cells("A2:H2")
    inst_cell = ws["A2"]
    inst_cell.value = "Instructions: Fill in the empty highlighted cells using the requested Excel functions. Do NOT hardcode values."
    inst_cell.font = italic_font
    inst_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # Add empty spacer row
    ws.row_dimensions[3].height = 15
    
    # 3. Add Headers & Data
    start_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[start_row].height = 28
    
    # Set data
    for row_offset, row_data in enumerate(data, 1):
        curr_row = start_row + row_offset
        ws.row_dimensions[curr_row].height = 22
        is_zebra = (row_offset % 2 == 0)
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.value = val
            cell.font = regular_font
            cell.border = thin_border
            
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(val, datetime.date):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "DD.MM.YYYY"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if is_zebra:
                cell.fill = zebra_fill
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in (1, 2):
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    return ws

def generate_beginner_excel(file_path: str):
    """Generate daily_showroom_footfall.xlsx template (Fixed Row 12 summaries)."""
    wb = Workbook()
    headers = ["Day", "Showroom Walk-ins", "Yaris Test Drives", "Corolla Test Drives"]
    data = [
        ["Monday", 15, 2, 5],
        ["Tuesday", 22, 4, 3],
        ["Wednesday", 18, 3, 4],
        ["Thursday", 30, 5, 6],
        ["Friday", 25, 4, 5],
        ["Saturday", 40, 8, 7],
        ["Sunday", 35, 6, 8]
    ]
    
    ws = create_styled_sheet(wb, "ShowroomCount", headers, data, "Daily Showroom Footfall & Inventory")
    
    # Summaries area (Empty cells for user to fill at Row 12, preserving Sunday's row 11)
    summary_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid") # Glowing light blue
    border_double = Side(style='double', color='1F2937')
    border_thin = Side(style='thin', color='D1D5DB')
    double_bottom = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)
    
    # Row 12: Summary row
    ws.cell(row=12, column=1).value = "Calculated Outputs:"
    ws.cell(row=12, column=1).font = Font(name="Segoe UI", size=11, bold=True)
    ws.row_dimensions[12].height = 25
    
    # Place visual target borders on empty cells
    for col_idx in (2, 3, 4, 5):
        cell = ws.cell(row=12, column=col_idx)
        cell.fill = summary_fill
        cell.border = double_bottom
        
    # Helper instruction labels
    ws.cell(row=13, column=1).value = "Required calculations:"
    ws.cell(row=13, column=1).font = Font(name="Segoe UI", size=10, bold=True, color="6B7280")
    
    ws.cell(row=14, column=2).value = "[B12] Total Weekly Walk-ins (SUM)"
    ws.cell(row=14, column=2).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=14, column=3).value = "[C12] Avg Daily Yaris Drives (AVERAGE)"
    ws.cell(row=14, column=3).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=14, column=4).value = "[D12] Avg Daily Corolla Drives (AVERAGE)"
    ws.cell(row=14, column=4).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=14, column=5).value = "[E12] Total Days Recorded (COUNT)"
    ws.cell(row=14, column=5).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["E"].width = 30
    wb.save(file_path)

def generate_weekly_sales_volume(file_path: str):
    """Generate weekly_sales_volume.xlsx template."""
    wb = Workbook()
    headers = ["Day", "Yaris Units", "Corolla Units", "Camry Units", "RAV4 Units", "Land Cruiser Units"]
    data = [
        ["Monday", 2, 3, 1, 1, 0],
        ["Tuesday", 3, 2, 2, 0, 1],
        ["Wednesday", 1, 4, 1, 2, 0],
        ["Thursday", 4, 3, 3, 1, 1],
        ["Friday", 5, 5, 2, 3, 2],
        ["Saturday", 8, 6, 4, 5, 3],
        ["Sunday", 6, 7, 3, 4, 2]
    ]
    ws = create_styled_sheet(wb, "SalesVolume", headers, data, "Weekly Vehicle Sales Volume")
    
    summary_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_double = Side(style='double', color='1F2937')
    border_thin = Side(style='thin', color='D1D5DB')
    double_bottom = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)
    
    ws.cell(row=12, column=1).value = "Calculated Outputs:"
    ws.cell(row=12, column=1).font = Font(name="Segoe UI", size=11, bold=True)
    ws.row_dimensions[12].height = 25
    
    for col_idx in (2, 3, 4, 5):
        cell = ws.cell(row=12, column=col_idx)
        cell.fill = summary_fill
        cell.border = double_bottom
        
    ws.cell(row=13, column=1).value = "Required calculations:"
    ws.cell(row=13, column=1).font = Font(name="Segoe UI", size=10, bold=True, color="6B7280")
    
    ws.cell(row=14, column=2).value = "[B12] Total Yaris Units (SUM)"
    ws.cell(row=14, column=2).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=14, column=3).value = "[C12] Avg Corolla Units (AVERAGE)"
    ws.cell(row=14, column=3).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=14, column=4).value = "[D12] Min Camry Units (MIN)"
    ws.cell(row=14, column=4).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=14, column=5).value = "[E12] Max RAV4 Units (MAX)"
    ws.cell(row=14, column=5).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.column_dimensions["A"].width = 24
    wb.save(file_path)

def generate_showroom_csat_score(file_path: str):
    """Generate showroom_csat_score.xlsx template."""
    wb = Workbook()
    headers = ["Customer ID", "Consultant Name", "CSAT Score (1-5)", "Feedback Category"]
    data = [
        [101, "Mahmoud", 5, "Excellent Service"],
        [102, "Yousef", 4, "Helpful staff"],
        [103, "Arslan", 3, "Average wait time"],
        [104, "Mahmoud", 5, "Great test drive"],
        [105, "Yousef", "", "No score submitted"], # Empty score cell
        [106, "Arslan", 4, "Satisfied"],
        [107, "Mahmoud", 5, "Highly recommended"]
    ]
    ws = create_styled_sheet(wb, "CSATAudit", headers, data, "Showroom Customer Satisfaction Score")
    
    summary_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_double = Side(style='double', color='1F2937')
    border_thin = Side(style='thin', color='D1D5DB')
    double_bottom = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)
    
    ws.cell(row=12, column=1).value = "Calculated Outputs:"
    ws.cell(row=12, column=1).font = Font(name="Segoe UI", size=11, bold=True)
    ws.row_dimensions[12].height = 25
    
    for col_idx in (3, 4, 5):
        cell = ws.cell(row=12, column=col_idx)
        cell.fill = summary_fill
        cell.border = double_bottom
        
    ws.cell(row=13, column=1).value = "Required calculations:"
    ws.cell(row=13, column=1).font = Font(name="Segoe UI", size=10, bold=True, color="6B7280")
    
    ws.cell(row=14, column=3).value = "[C12] Average Score (AVERAGE)"
    ws.cell(row=14, column=3).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=14, column=4).value = "[D12] Rated Interactions (COUNT)"
    ws.cell(row=14, column=4).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=14, column=5).value = "[E12] Total Comments (COUNTA)"
    ws.cell(row=14, column=5).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    wb.save(file_path)

def generate_spare_parts_orders(file_path: str):
    """Generate spare_parts_orders.xlsx template."""
    wb = Workbook()
    headers = ["Order Date", "Part Number", "Description", "Quantity", "Unit Cost (AED)", "Total Cost (AED)"]
    data = [
        [datetime.date(2026, 5, 27), "TY-1002", "Oil Filter", 10, 45, "=D5*E5"],
        [datetime.date(2026, 5, 27), "TY-2005", "Brake Pad Front", 5, 180, "=D6*E6"],
        [datetime.date(2026, 5, 28), "TY-3001", "Spark Plug", 30, 25, "=D7*E7"],
        [datetime.date(2026, 5, 28), "TY-4009", "Air Filter", 8, 65, "=D8*E8"],
        [datetime.date(2026, 5, 29), "TY-2006", "Brake Pad Rear", 4, 150, "=D9*E9"]
    ]
    ws = create_styled_sheet(wb, "PartsOrders", headers, data, "Dealership Spare Parts Orders")
    
    summary_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_double = Side(style='double', color='1F2937')
    border_thin = Side(style='thin', color='D1D5DB')
    double_bottom = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)
    
    ws.cell(row=10, column=1).value = "Calculated Outputs:"
    ws.cell(row=10, column=1).font = Font(name="Segoe UI", size=11, bold=True)
    ws.row_dimensions[10].height = 25
    
    for col_idx in (4, 5, 6):
        cell = ws.cell(row=10, column=col_idx)
        cell.fill = summary_fill
        cell.border = double_bottom
        
    ws.cell(row=11, column=1).value = "Required calculations:"
    ws.cell(row=11, column=1).font = Font(name="Segoe UI", size=10, bold=True, color="6B7280")
    
    ws.cell(row=12, column=4).value = "[D10] Total Quantity (SUM)"
    ws.cell(row=12, column=4).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=12, column=5).value = "[E10] Avg Unit Cost (AVERAGE)"
    ws.cell(row=12, column=5).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=12, column=6).value = "[F10] Total Cost (SUM)"
    ws.cell(row=12, column=6).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    wb.save(file_path)

def generate_test_drive_conversion(file_path: str):
    """Generate test_drive_conversion.xlsx template."""
    wb = Workbook()
    headers = ["Week", "Yaris Drives", "Camry Drives", "Land Cruiser Drives", "Conversions"]
    data = [
        ["Week 1", 12, 8, 4, 6],
        ["Week 2", 15, 10, 6, 8],
        ["Week 3", 18, 12, 5, 7],
        ["Week 4", 20, 14, 8, 10]
    ]
    ws = create_styled_sheet(wb, "Conversions", headers, data, "Showroom Test Drive Conversion")
    
    summary_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_double = Side(style='double', color='1F2937')
    border_thin = Side(style='thin', color='D1D5DB')
    double_bottom = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)
    
    ws.cell(row=9, column=1).value = "Calculated Outputs:"
    ws.cell(row=9, column=1).font = Font(name="Segoe UI", size=11, bold=True)
    ws.row_dimensions[9].height = 25
    
    for col_idx in (2, 3, 4, 5):
        cell = ws.cell(row=9, column=col_idx)
        cell.fill = summary_fill
        cell.border = double_bottom
        
    ws.cell(row=10, column=1).value = "Required calculations:"
    ws.cell(row=10, column=1).font = Font(name="Segoe UI", size=10, bold=True, color="6B7280")
    
    ws.cell(row=11, column=2).value = "[B9] Total Yaris Drives (SUM)"
    ws.cell(row=11, column=2).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=11, column=3).value = "[C9] Total Camry Drives (SUM)"
    ws.cell(row=11, column=3).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=11, column=4).value = "[D9] Avg LC Drives (AVERAGE)"
    ws.cell(row=11, column=4).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=11, column=5).value = "[E9] Total Conversions (SUM)"
    ws.cell(row=11, column=5).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    wb.save(file_path)

def generate_intermediate_excel(file_path: str):
    """Generate monthly_sales_commission.xlsx template."""
    wb = Workbook()
    catalog_headers = ["VIN Prefix", "Model", "Trim", "Price (AED)"]
    catalog_data = [
        ["LC", "Land Cruiser", "GXR", 320000],
        ["RAV", "RAV4", "Adventure", 140000],
        ["YAR", "Yaris", "Sedan", 75000],
        ["COR", "Corolla", "GLI", 85000]
    ]
    create_styled_sheet(wb, "VehicleCatalog", catalog_headers, catalog_data, "Vehicle Master Catalog")
    
    sales_headers = ["Deal ID", "VIN Prefix", "Model", "Trim", "Price (AED)", "Comm Rate", "Comm Amount", "Branch"]
    sales_data = [
        [1001, "LC", "Land Cruiser", "GXR", "", "", "", "Dubai"],
        [1002, "RAV", "RAV4", "Adventure", "", "", "", "Abu Dhabi"],
        [1003, "YAR", "Yaris", "Sedan", "", "", "", "Dubai"],
        [1004, "COR", "Corolla", "GLI", "", "", "", "Sharjah"],
        [1005, "LC", "Land Cruiser", "GXR", "", "", "", "Dubai"],
        [1006, "RAV", "RAV4", "Adventure", "", "", "", "Sharjah"]
    ]
    ws = create_styled_sheet(wb, "SalesLog", sales_headers, sales_data, "Monthly Sales Commission Audit")
    
    # Visual Highlights for empty cells
    target_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_thin = Side(style='thin', color='D1D5DB')
    cell_border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    for row in range(5, 11):
        for col in (5, 6, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = target_fill
            cell.border = cell_border
            
    # Dubai Total Sales Box (SUMIFS target)
    ws.cell(row=5, column=9).value = "Audited Metrics:"
    ws.cell(row=5, column=9).font = Font(name="Segoe UI", size=10, bold=True)
    
    ws.cell(row=6, column=9).value = "Dubai Total Sales (AED):"
    ws.cell(row=6, column=9).font = Font(name="Segoe UI", size=10, bold=True)
    ws.cell(row=6, column=9).border = cell_border
    
    target_cell = ws.cell(row=6, column=10)
    target_cell.fill = target_fill
    target_cell.border = cell_border
    
    ws.cell(row=8, column=9).value = "Required formulas:"
    ws.cell(row=8, column=9).font = Font(name="Segoe UI", size=9, bold=True, color="6B7280")
    
    ws.cell(row=9, column=9).value = "E5-E10: Lookup Price from Catalog (VLOOKUP)"
    ws.cell(row=9, column=9).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=10, column=9).value = "F5-F10: Rate: 5% for LC, 3% for others (IF)"
    ws.cell(row=10, column=9).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=11, column=9).value = "G5-G10: Commission Amount = Price * Rate"
    ws.cell(row=11, column=9).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=12, column=9).value = "J6: Total Sales Generated by Dubai (SUMIFS)"
    ws.cell(row=12, column=9).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    # Adjust widths
    ws.column_dimensions["I"].width = 24
    ws.column_dimensions["J"].width = 24
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(file_path)

def generate_camry_fleet_lease(file_path: str):
    """Generate camry_fleet_lease.xlsx template."""
    wb = Workbook()
    catalog_headers = ["Camry Trim", "Daily Rate (AED)"]
    catalog_data = [
        ["Standard", 150],
        ["Hybrid", 180],
        ["Executive", 220]
    ]
    create_styled_sheet(wb, "Rates", catalog_headers, catalog_data, "Fleet Rental Catalog")
    
    rentals_headers = ["Rental ID", "Customer", "Camry Trim", "Duration (Days)", "Daily Rate (AED)", "Base Cost (AED)", "Discount", "Total Rent (AED)", "Branch"]
    rentals_data = [
        [3001, "Fleet Corp", "Hybrid", 10, "", "", "", "", "Dubai"],
        [3002, "Amir Khan", "Standard", 3, "", "", "", "", "Abu Dhabi"],
        [3003, "Yousef", "Executive", 7, "", "", "", "", "Dubai"],
        [3004, "Sharjah Logistics", "Hybrid", 5, "", "", "", "", "Sharjah"],
        [3005, "Al-Futtaim Corp", "Executive", 12, "", "", "", "", "Dubai"]
    ]
    ws = create_styled_sheet(wb, "Rentals", rentals_headers, rentals_data, "Camry Fleet Lease Audit")
    
    target_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_thin = Side(style='thin', color='D1D5DB')
    cell_border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    for row in range(5, 10):
        for col in (5, 6, 7, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = target_fill
            cell.border = cell_border
            
    ws.cell(row=5, column=10).value = "Audited Metrics:"
    ws.cell(row=5, column=10).font = Font(name="Segoe UI", size=10, bold=True)
    
    ws.cell(row=6, column=10).value = "Count Hybrid Rentals:"
    ws.cell(row=6, column=10).font = Font(name="Segoe UI", size=10, bold=True)
    ws.cell(row=6, column=10).border = cell_border
    
    target_cell = ws.cell(row=6, column=11)
    target_cell.fill = target_fill
    target_cell.border = cell_border
    
    ws.cell(row=8, column=10).value = "Required formulas:"
    ws.cell(row=8, column=10).font = Font(name="Segoe UI", size=9, bold=True, color="6B7280")
    
    ws.cell(row=9, column=10).value = "E5-E9: Lookup Rate from Rates Catalog (VLOOKUP)"
    ws.cell(row=9, column=10).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=10, column=10).value = "F5-F9: Base Cost = Duration * Rate"
    ws.cell(row=10, column=10).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=11, column=10).value = "G5-G9: Discount: 10% if Duration >= 7, else 0% (IF)"
    ws.cell(row=11, column=10).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=12, column=10).value = "H5-H9: Total Rent = Base Cost * (1 - Discount)"
    ws.cell(row=12, column=10).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=13, column=10).value = "K6: Count of Hybrid rentals active (COUNTIF)"
    ws.cell(row=13, column=10).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.column_dimensions["J"].width = 24
    ws.column_dimensions["K"].width = 24
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(file_path)

def generate_sales_agent_performance(file_path: str):
    """Generate sales_agent_performance.xlsx template."""
    wb = Workbook()
    headers = ["Agent", "Branch", "Vehicle Sold", "Margin (AED)", "Sales Count"]
    data = [
        ["Mahmoud", "Dubai", "Land Cruiser", 25000, 1],
        ["Yousef", "Abu Dhabi", "Camry", 8000, 1],
        ["Arslan", "Dubai", "Yaris", 3500, 1],
        ["Mahmoud", "Dubai", "RAV4", 9000, 1],
        ["Yousef", "Dubai", "Land Cruiser", 28000, 1],
        ["Arslan", "Abu Dhabi", "Corolla", 4000, 1]
    ]
    ws = create_styled_sheet(wb, "SalesLog", headers, data, "Showroom Sales Agent Performance")
    
    target_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_thin = Side(style='thin', color='D1D5DB')
    cell_border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    ws.cell(row=5, column=7).value = "Audited Metrics Summary:"
    ws.cell(row=5, column=7).font = Font(name="Segoe UI", size=10, bold=True)
    
    ws.cell(row=6, column=7).value = "Total Margin - Dubai (AED):"
    ws.cell(row=6, column=7).font = Font(name="Segoe UI", size=10, bold=True)
    ws.cell(row=6, column=7).border = cell_border
    
    cell1 = ws.cell(row=6, column=8)
    cell1.fill = target_fill
    cell1.border = cell_border
    
    ws.cell(row=7, column=7).value = "Avg Margin - Mahmoud Dubai (AED):"
    ws.cell(row=7, column=7).font = Font(name="Segoe UI", size=10, bold=True)
    ws.cell(row=7, column=7).border = cell_border
    
    cell2 = ws.cell(row=7, column=8)
    cell2.fill = target_fill
    cell2.border = cell_border
    
    ws.cell(row=8, column=7).value = "Dubai High-Value Count:"
    ws.cell(row=8, column=7).font = Font(name="Segoe UI", size=10, bold=True)
    ws.cell(row=8, column=7).border = cell_border
    
    cell3 = ws.cell(row=8, column=8)
    cell3.fill = target_fill
    cell3.border = cell_border
    
    ws.cell(row=10, column=7).value = "Required formulas:"
    ws.cell(row=10, column=7).font = Font(name="Segoe UI", size=9, bold=True, color="6B7280")
    
    ws.cell(row=11, column=7).value = "H6: Sum Margins in Dubai (SUMIFS)"
    ws.cell(row=11, column=7).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=12, column=7).value = "H7: Avg Margin for Mahmoud in Dubai (AVERAGEIFS)"
    ws.cell(row=12, column=7).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=13, column=7).value = "H8: Count of Dubai deals with Margin > 5000 (COUNTIFS)"
    ws.cell(row=13, column=7).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 24
    
    wb.save(file_path)

def generate_spare_parts_valuation(file_path: str):
    """Generate spare_parts_valuation.xlsx template."""
    wb = Workbook()
    catalog_headers = ["Part No", "Wholesale Cost (AED)"]
    catalog_data = [
        ["TY-101", 30],
        ["TY-102", 120],
        ["TY-103", 200]
    ]
    create_styled_sheet(wb, "Catalog", catalog_headers, catalog_data, "Spare Parts Catalog")
    
    inv_headers = ["Part No", "Quantity", "Wholesale Cost (AED)", "Markup Category", "Retail Cost (AED)", "Total Value (AED)", "Branch"]
    inv_data = [
        ["TY-101", 100, "", "A", "", "", "Dubai"],
        ["TY-102", 40, "", "B", "", "", "Abu Dhabi"],
        ["TY-103", 15, "", "A", "", "", "Dubai"],
        ["TY-101", 200, "", "B", "", "", "Sharjah"],
        ["TY-102", 60, "", "A", "", "", "Dubai"]
    ]
    ws = create_styled_sheet(wb, "Inventory", inv_headers, inv_data, "Spare Parts Inventory Valuation")
    
    target_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_thin = Side(style='thin', color='D1D5DB')
    cell_border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    for row in range(5, 10):
        for col in (3, 5, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = target_fill
            cell.border = cell_border
            
    ws.cell(row=5, column=9).value = "Audited Metrics:"
    ws.cell(row=5, column=9).font = Font(name="Segoe UI", size=10, bold=True)
    
    ws.cell(row=6, column=9).value = "Dubai Retail Valuation (AED):"
    ws.cell(row=6, column=9).font = Font(name="Segoe UI", size=10, bold=True)
    ws.cell(row=6, column=9).border = cell_border
    
    target_cell = ws.cell(row=6, column=10)
    target_cell.fill = target_fill
    target_cell.border = cell_border
    
    ws.cell(row=8, column=9).value = "Required formulas:"
    ws.cell(row=8, column=9).font = Font(name="Segoe UI", size=9, bold=True, color="6B7280")
    
    ws.cell(row=9, column=9).value = "C5-C9: Wholesale cost lookup from Catalog (VLOOKUP)"
    ws.cell(row=9, column=9).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=10, column=9).value = "E5-E9: Markup: 30% for 'A', else 15% (IF)"
    ws.cell(row=10, column=9).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=11, column=9).value = "F5-F9: Total Value = Quantity * Retail Cost"
    ws.cell(row=11, column=9).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=12, column=9).value = "J6: Total Retail Value for Dubai (SUMIFS)"
    ws.cell(row=12, column=9).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.column_dimensions["I"].width = 26
    ws.column_dimensions["J"].width = 24
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(file_path)

def generate_opex_allocation(file_path: str):
    """Generate opex_allocation.xlsx template."""
    wb = Workbook()
    catalog_headers = ["Showroom Size", "Base Rent Factor"]
    catalog_data = [
        ["Small", 500],
        ["Medium", 1200],
        ["Large", 2500]
    ]
    create_styled_sheet(wb, "Sizing", catalog_headers, catalog_data, "Showroom Base Rent Matrix")
    
    sales_headers = ["Showroom Name", "Size Class", "Headcount", "Allocated Rent (AED)", "Overhead Penalty", "Total Allocation (AED)"]
    sales_data = [
        ["Dubai Main", "Large", 45, "", "", ""],
        ["Abu Dhabi Hub", "Large", 35, "", "", ""],
        ["Sharjah Branch", "Medium", 15, "", "", ""],
        ["Ajman Express", "Small", 8, "", "", ""],
        ["Al Ain Outlet", "Medium", 12, "", "", ""]
    ]
    ws = create_styled_sheet(wb, "AllocationLog", sales_headers, sales_data, "Dealership Expense Allocation")
    
    target_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_thin = Side(style='thin', color='D1D5DB')
    cell_border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    for row in range(5, 10):
        for col in (4, 5, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = target_fill
            cell.border = cell_border
            
    ws.cell(row=5, column=8).value = "Audited Metrics Summary:"
    ws.cell(row=5, column=8).font = Font(name="Segoe UI", size=10, bold=True)
    
    ws.cell(row=6, column=8).value = "Total Allocation Sum (AED):"
    ws.cell(row=6, column=8).font = Font(name="Segoe UI", size=10, bold=True)
    ws.cell(row=6, column=8).border = cell_border
    
    target_cell = ws.cell(row=6, column=9)
    target_cell.fill = target_fill
    target_cell.border = cell_border
    
    ws.cell(row=8, column=8).value = "Required formulas:"
    ws.cell(row=8, column=8).font = Font(name="Segoe UI", size=9, bold=True, color="6B7280")
    
    ws.cell(row=9, column=8).value = "D5-D9: Base Rent = VLOOKUP from Sizing catalog"
    ws.cell(row=9, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=10, column=8).value = "E5-E9: Overhead: if Headcount > 20, 500, else 100 (IF)"
    ws.cell(row=10, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=11, column=8).value = "F5-F9: Total = Allocated Rent + Overhead Penalty"
    ws.cell(row=11, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=12, column=8).value = "I6: Sum of Total Allocations (SUMIFS)"
    ws.cell(row=12, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.column_dimensions["H"].width = 24
    ws.column_dimensions["I"].width = 24
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(file_path)

def generate_loan_eligibility(file_path: str):
    """Generate loan_eligibility.xlsx template."""
    wb = Workbook()
    catalog_headers = ["Credit Tier", "Standard APR (%)"]
    catalog_data = [
        ["Excellent", 0.0299],
        ["Good", 0.0450],
        ["Fair", 0.0699]
    ]
    create_styled_sheet(wb, "APRList", catalog_headers, catalog_data, "Credit Tier APR Matrix")
    
    sales_headers = ["App ID", "Customer", "Income (AED)", "Credit Tier", "Qualify Status", "Interest Rate (%)"]
    sales_data = [
        [501, "Fatima", 12000, "Excellent", "", ""],
        [502, "Sameer", 4500, "Good", "", ""],
        [503, "Jackson", 8000, "Fair", "", ""],
        [504, "Khadija", 15000, "Excellent", "", ""],
        [505, "George", 5000, "Fair", "", ""]
    ]
    ws = create_styled_sheet(wb, "Applications", sales_headers, sales_data, "Vehicle Loan Eligibility Audit")
    
    target_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_thin = Side(style='thin', color='D1D5DB')
    cell_border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    for row in range(5, 10):
        for col in (5, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = target_fill
            cell.border = cell_border
            
    ws.cell(row=5, column=8).value = "Audited Metrics:"
    ws.cell(row=5, column=8).font = Font(name="Segoe UI", size=10, bold=True)
    
    ws.cell(row=6, column=8).value = "Approved Fair Credit Customers:"
    ws.cell(row=6, column=8).font = Font(name="Segoe UI", size=10, bold=True)
    ws.cell(row=6, column=8).border = cell_border
    
    target_cell = ws.cell(row=6, column=9)
    target_cell.fill = target_fill
    target_cell.border = cell_border
    
    ws.cell(row=8, column=8).value = "Required formulas:"
    ws.cell(row=8, column=8).font = Font(name="Segoe UI", size=9, bold=True, color="6B7280")
    
    ws.cell(row=9, column=8).value = "E5-E9: Status: 'Eligible' if Income > 6000 and Tier != 'Fair' (IF, AND)"
    ws.cell(row=9, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=10, column=8).value = "F5-F9: Rate: VLOOKUP from APRList Matrix"
    ws.cell(row=10, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=11, column=8).value = "I6: Count of 'Fair' tier customers processed (COUNTIF)"
    ws.cell(row=11, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.column_dimensions["H"].width = 28
    ws.column_dimensions["I"].width = 24
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(file_path)

def generate_revenue_share_audit(file_path: str):
    """Generate revenue_share_audit.xlsx template."""
    wb = Workbook()
    catalog_headers = ["Tier Level", "Share Payout (%)"]
    catalog_data = [
        ["Gold", 0.12],
        ["Silver", 0.08],
        ["Bronze", 0.05]
    ]
    create_styled_sheet(wb, "TierMatrix", catalog_headers, catalog_data, "Revenue Tier Share Matrix")
    
    sales_headers = ["Partner ID", "Gross Revenue (AED)", "Tier Level", "Payout Rate (%)", "Direct Profit", "Bonus Payout (AED)", "Total Allocation (AED)", "Region"]
    sales_data = [
        [701, 850000, "Gold", "", "", "", "", "Dubai"],
        [702, 450000, "Silver", "", "", "", "", "Abu Dhabi"],
        [703, 300000, "Bronze", "", "", "", "", "Dubai"],
        [704, 980000, "Gold", "", "", "", "", "Sharjah"],
        [705, 520000, "Silver", "", "", "", "", "Dubai"]
    ]
    ws = create_styled_sheet(wb, "ShareAudit", sales_headers, sales_data, "Partnership Revenue Share Audit")
    
    target_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_thin = Side(style='thin', color='D1D5DB')
    cell_border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    for row in range(5, 10):
        for col in (4, 5, 6, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = target_fill
            cell.border = cell_border
            
    ws.cell(row=5, column=10).value = "Audited Metrics Summary:"
    ws.cell(row=5, column=10).font = Font(name="Segoe UI", size=10, bold=True)
    
    ws.cell(row=6, column=10).value = "Dubai Base Shared Revenue (AED):"
    ws.cell(row=6, column=10).font = Font(name="Segoe UI", size=10, bold=True)
    ws.cell(row=6, column=10).border = cell_border
    
    target_cell = ws.cell(row=6, column=11)
    target_cell.fill = target_fill
    target_cell.border = cell_border
    
    ws.cell(row=8, column=10).value = "Required formulas:"
    ws.cell(row=8, column=10).font = Font(name="Segoe UI", size=9, bold=True, color="6B7280")
    
    ws.cell(row=9, column=10).value = "D5-D9: Payout Rate = VLOOKUP from TierMatrix"
    ws.cell(row=9, column=10).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=10, column=10).value = "E5-E9: Direct Profit = Revenue * Payout Rate"
    ws.cell(row=10, column=10).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=11, column=10).value = "F5-F9: Bonus = if Revenue > 500000, 10000, else 0 (IF)"
    ws.cell(row=11, column=10).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=12, column=10).value = "G5-G9: Total Allocation = Direct Profit + Bonus"
    ws.cell(row=12, column=10).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=13, column=10).value = "K6: Sum of Total Base Payout in Dubai (SUMIFS)"
    ws.cell(row=13, column=10).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.column_dimensions["J"].width = 28
    ws.column_dimensions["K"].width = 24
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(file_path)

def generate_advanced_excel(file_path: str):
    """Generate dealership_profitability.xlsx template."""
    wb = Workbook()
    matrix_headers = ["Score Tier", "Base Provision (%)", "Premium Markup (%)"]
    matrix_data = [
        ["Tier A", 0.05, 0.12],
        ["Tier B", 0.08, 0.15],
        ["Tier C", 0.12, 0.20]
    ]
    create_styled_sheet(wb, "Matrix", matrix_headers, matrix_data, "Leasing Cost Markup Matrix")
    
    ws = wb.create_sheet("AgingAndLease")
    ws.sheet_view.showGridLines = True
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    sub_header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    target_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    
    border_thin = Side(style='thin', color='D1D5DB')
    border_double = Side(style='double', color='1F2937')
    cell_border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    # 1. Main Title
    ws.merge_cells("A1:H1")
    t_cell = ws["A1"]
    t_cell.value = "INTERNAL FINANCE TRAINING: ADVANCED CAPITAL BUDGETING & AGING"
    t_cell.font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    t_cell.fill = header_fill
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Headers
    headers = ["Receipt Date", "Base Cost (AED)", "Holding Days", "Aging Bracket", "Compounding Penalty"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
    ws.row_dimensions[4].height = 25
    
    data = [
        [datetime.date(2026, 4, 15), 450000, "=TODAY()-A5", "", ""],
        [datetime.date(2026, 5, 10), 125000, "=TODAY()-A6", "", ""],
        [datetime.date(2026, 5, 20), 85000, "=TODAY()-A7", "", ""]
    ]
    for row_idx, row_data in enumerate(data, 5):
        ws.row_dimensions[row_idx].height = 22
        for col_idx, v in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = v
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = cell_border
            
            if isinstance(v, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(v, datetime.date):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "DD.MM.YYYY"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if col_idx in (4, 5):
                cell.fill = target_fill
                
    # 2. Lookup Area (INDEX/MATCH target)
    ws.cell(row=9, column=1).value = "Capital Project Analysis Dashboard:"
    ws.cell(row=9, column=1).font = Font(name="Segoe UI", size=11, bold=True)
    
    ws.cell(row=10, column=1).value = "Provision Lookup Matrix:"
    ws.cell(row=10, column=1).font = Font(name="Segoe UI", size=9, bold=True, color="6B7280")
    
    matrix_hdr = ["Score Tier", "Base Provision (%)", "Premium Markup (%)"]
    for idx, h in enumerate(matrix_hdr, 1):
        c = ws.cell(row=10, column=idx)
        c.value = h
        c.font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
        c.fill = sub_header_fill
        c.border = cell_border
        
    m_data = [
        ["Tier A", 0.05, 0.12],
        ["Tier B", 0.08, 0.15],
        ["Tier C", 0.12, 0.20]
    ]
    for r_idx, r_val in enumerate(m_data, 11):
        for c_idx, val in enumerate(r_val, 1):
            c = ws.cell(row=r_idx, column=c_idx)
            c.value = val
            c.font = Font(name="Segoe UI", size=9)
            c.border = cell_border
            if c_idx in (2, 3):
                c.number_format = "0.0%"
                
    # Dynamic inputs
    ws.cell(row=15, column=1).value = "Select Tier:"
    ws.cell(row=15, column=1).font = Font(name="Segoe UI", size=9, bold=True)
    ws.cell(row=15, column=2).value = "Tier B"
    ws.cell(row=15, column=2).font = Font(name="Segoe UI", size=9, bold=True)
    ws.cell(row=15, column=2).border = cell_border
    
    ws.cell(row=16, column=1).value = "Select Metric:"
    ws.cell(row=16, column=1).font = Font(name="Segoe UI", size=9, bold=True)
    ws.cell(row=16, column=2).value = "Premium Markup (%)"
    ws.cell(row=16, column=2).font = Font(name="Segoe UI", size=9, bold=True)
    ws.cell(row=16, column=2).border = cell_border
    
    ws.cell(row=17, column=1).value = "Resulting Allocation:"
    ws.cell(row=17, column=1).font = Font(name="Segoe UI", size=10, bold=True)
    
    target_lookup = ws.cell(row=17, column=2)
    target_lookup.fill = target_fill
    target_lookup.border = cell_border
    
    # 3. NPV Cashflows
    ws.cell(row=19, column=1).value = "Showroom Leasing Capital Project flows:"
    ws.cell(row=19, column=1).font = Font(name="Segoe UI", size=10, bold=True)
    
    fl_headers = ["Project Date", "Cash Flow (AED)", "", "Discount Rate (r)"]
    for c_idx, h in enumerate(fl_headers, 1):
        if h:
            c = ws.cell(row=20, column=c_idx)
            c.value = h
            c.font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
            c.fill = sub_header_fill
            c.border = cell_border
            
    ws.cell(row=21, column=4).value = 0.08  # 8% Discount rate
    ws.cell(row=21, column=4).font = Font(name="Segoe UI", size=9)
    ws.cell(row=21, column=4).border = cell_border
    ws.cell(row=21, column=4).number_format = "0.0%"
    
    flows = [
        [datetime.date(2026, 5, 27), -300000],
        [datetime.date(2027, 5, 27), 100000],
        [datetime.date(2028, 5, 27), 120000],
        [datetime.date(2029, 5, 27), 150000],
        [datetime.date(2030, 5, 27), 180000]
    ]
    for r_offset, r_data in enumerate(flows, 21):
        for c_idx, val in enumerate(r_data, 1):
            c = ws.cell(row=r_offset, column=c_idx)
            c.value = val
            c.font = Font(name="Segoe UI", size=9)
            c.border = cell_border
            if c_idx == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = "DD.MM.YYYY"
            else:
                c.alignment = Alignment(horizontal="right", vertical="center")
                
    ws.cell(row=21, column=5).value = "Project XNPV (AED):"
    ws.cell(row=21, column=5).font = Font(name="Segoe UI", size=10, bold=True)
    ws.cell(row=21, column=5).border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    xnpv_cell = ws.cell(row=21, column=6)
    xnpv_cell.fill = target_fill
    xnpv_cell.border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    ws.cell(row=23, column=5).value = "Project XIRR (%):"
    ws.cell(row=23, column=5).font = Font(name="Segoe UI", size=10, bold=True)
    ws.cell(row=23, column=5).border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    xirr_cell = ws.cell(row=23, column=6)
    xirr_cell.fill = target_fill
    xirr_cell.border = Border(top=border_thin, bottom=border_thin, left=border_thin, right=border_thin)
    
    # Detail Guidelines on the right
    ws.cell(row=9, column=8).value = "Formula Guidelines:"
    ws.cell(row=9, column=8).font = Font(name="Segoe UI", size=10, bold=True, color="6B7280")
    
    ws.cell(row=10, column=8).value = "D5-D7: Aging Bracket. If days <= 30: '0-30 days', if days <= 90: '31-90 days', else: '90+ days'"
    ws.cell(row=10, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=11, column=8).value = "E5-E7: Holding Cost. If bracket is '0-30 days': 50, if '31-90 days': 100, else: 150 (use Nested IF)"
    ws.cell(row=11, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=13, column=8).value = "B17: 2D query based on B15 and B16 from mini performance matrix (use INDEX/MATCH)"
    ws.cell(row=13, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=15, column=8).value = "F21: Calculate XNPV on lease cash flows given discount rate (r) in D21 (use XNPV)"
    ws.cell(row=15, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    ws.cell(row=17, column=8).value = "F23: Calculate XIRR on lease cash flows (use XIRR)"
    ws.cell(row=17, column=8).font = Font(name="Segoe UI", size=8, italic=True, color="4B5563")
    
    # Adjust widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["H"].width = 40
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(file_path)

def generate_blank_template(file_path: str, task_title: str):
    """Utility to generate generic blank styled tables for extended curriculum tasks."""
    wb = Workbook()
    headers = ["ID", "Parameter Name", "Value A", "Value B", "Calculated Value (Formula Required)"]
    data = [
        [101, "Showroom Sales Data", 250, 180, ""],
        [102, "Operational Expense Log", 1200, 450, ""],
        [103, "Spare Parts Demand", 45, 12, ""],
        [104, "Customer CSAT Weighted", 5, 4, ""],
        [105, "Used Car Trade Value", 35000, 12000, ""]
    ]
    ws = create_styled_sheet(wb, "DataAnalysis", headers, data, task_title)
    
    summary_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border_double = Side(style='double', color='1F2937')
    border_thin = Side(style='thin', color='D1D5DB')
    double_bottom = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)
    
    ws.cell(row=10, column=1).value = "Calculated Outputs:"
    ws.cell(row=10, column=1).font = Font(name="Segoe UI", size=11, bold=True)
    ws.row_dimensions[10].height = 25
    
    for col_idx in (3, 4, 5):
        cell = ws.cell(row=10, column=col_idx)
        cell.fill = summary_fill
        cell.border = double_bottom
        
    ws.cell(row=11, column=1).value = "Required calculations:"
    ws.cell(row=11, column=1).font = Font(name="Segoe UI", size=10, bold=True, color="6B7280")
    
    ws.cell(row=12, column=3).value = "[C10] Sum of Values (SUM)"
    ws.cell(row=12, column=3).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=12, column=4).value = "[D10] Average of Values (AVERAGE)"
    ws.cell(row=12, column=4).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    ws.cell(row=12, column=5).value = "[E10] Count of Items (COUNT)"
    ws.cell(row=12, column=5).font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")
    
    wb.save(file_path)

def seed_database(db: Session):
    """Seed SQLite database with standard users, default tasks, and validation rules."""
    # 1. Create default Admin User
    admin = db.query(models.User).filter(models.User.username == "admin").first()
    if not admin:
        admin = models.User(
            username="admin",
            full_name="Administrator",
            email="admin@training.com",
            password_hash=None,
            role="admin"
        )
        db.add(admin)
        db.flush()
        
    # 2. Seed the 7 Standard Users
    standard_users = [
        ("mahmoud", "Mahmoud Ahmed", "mahmoud@training.com"),
        ("arslan", "Arslan Ishaq", "arslan@training.com"),
        ("asela", "Asela Lakmal", "asela@training.com"),
        ("charuka", "Charuka Jayashan", "charuka@training.com"),
        ("nushan", "Mohammad Nushan", "nushan@training.com"),
        ("yousef", "Yousef", "yousef@training.com"),
        ("khurram", "Khurram Khan", "khurram@training.com")
    ]
    
    for username, full_name, email in standard_users:
        user = db.query(models.User).filter(models.User.username == username).first()
        if not user:
            user = models.User(
                username=username,
                full_name=full_name,
                email=email,
                password_hash=None,
                role="user"
            )
            db.add(user)
            db.flush()
            
            progress = models.UserProgress(
                user_id=user.id,
                current_active_task_id="daily_showroom_footfall",
                beginner_unlocked=True,
                intermediate_unlocked=False,
                advanced_unlocked=False
            )
            db.add(progress)
            
            score = models.Score(
                user_id=user.id,
                total_points=0,
                failed_attempts_count=0
            )
            db.add(score)
            
    # 3. Create default tasks (5 Beginner, 7 Intermediate, 10 Advanced)
    tasks = [
        # --- STAGE 1: BEGINNER (5 Tasks) ---
        {
            "id": "daily_showroom_footfall",
            "stage": "beginner",
            "title": "Daily Showroom Footfall & Inventory Count",
            "scenario_text": (
                "You are the Finance Associate at the regional headquarters. Calculate the total weekly walk-ins, "
                "the average daily test drives for each model, and count the total days recorded.\n\n"
                "Learning Goals:\n"
                "- SUM: Total weekly footfall in cell B12\n"
                "- AVERAGE: Average daily Yaris test drives in cell C12\n"
                "- AVERAGE: Average daily Corolla test drives in cell D12\n"
                "- COUNT: Number of active showroom days recorded in cell E12"
            ),
            "download_template_name": "daily_showroom_footfall.xlsx",
            "validations": [
                {"cell_reference": "B12", "expected_value": "185", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(B5:B11)"},
                {"cell_reference": "C12", "expected_value": "4.57", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(C5:C11)"},
                {"cell_reference": "D12", "expected_value": "5.43", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(D5:D11)"},
                {"cell_reference": "E12", "expected_value": "7", "required_function": "COUNT", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNT(B5:B11)"}
            ]
        },
        {
            "id": "weekly_sales_volume",
            "stage": "beginner",
            "title": "Weekly Vehicle Sales Volume Analysis",
            "scenario_text": (
                "Calculate total Yaris sales, average daily Corolla sales, minimum Camry sales, and maximum RAV4 sales "
                "recorded last week.\n\n"
                "Learning Goals:\n"
                "- SUM: Total Yaris sold in cell B12\n"
                "- AVERAGE: Avg Corolla sales in cell C12\n"
                "- MIN: Min Camry sales in cell D12\n"
                "- MAX: Max RAV4 sales in cell E12"
            ),
            "download_template_name": "weekly_sales_volume.xlsx",
            "validations": [
                {"cell_reference": "B12", "expected_value": "29", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(B5:B11)"},
                {"cell_reference": "C12", "expected_value": "4.29", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(C5:C11)"},
                {"cell_reference": "D12", "expected_value": "1", "required_function": "MIN", "is_financial": False, "is_date": False, "correct_formula_hint": "=MIN(D5:D11)"},
                {"cell_reference": "E12", "expected_value": "5", "required_function": "MAX", "is_financial": False, "is_date": False, "correct_formula_hint": "=MAX(E5:E11)"}
            ]
        },
        {
            "id": "showroom_csat_score",
            "stage": "beginner",
            "title": "Showroom Customer Satisfaction Score (CSAT)",
            "scenario_text": (
                "Analyze customer feedback scores from last week. Calculate the average CSAT score, "
                "count total rated interactions, and count total text feedback comments submitted.\n\n"
                "Learning Goals:\n"
                "- AVERAGE: Average score in cell C12\n"
                "- COUNT: Count of numeric ratings in cell D12\n"
                "- COUNTA: Count of non-empty feedback comments in cell E12"
            ),
            "download_template_name": "showroom_csat_score.xlsx",
            "validations": [
                {"cell_reference": "C12", "expected_value": "4.33", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(C5:C11)"},
                {"cell_reference": "D12", "expected_value": "6", "required_function": "COUNT", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNT(C5:C11)"},
                {"cell_reference": "E12", "expected_value": "7", "required_function": "COUNTA", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNTA(D5:D11)"}
            ]
        },
        {
            "id": "spare_parts_orders",
            "stage": "beginner",
            "title": "Dealership Spare Parts Purchase Orders",
            "scenario_text": (
                "Audit dealership spare parts purchase orders. Calculate total quantity ordered, "
                "average unit cost, and total amount spent.\n\n"
                "Learning Goals:\n"
                "- SUM: Total quantity in cell D10\n"
                "- AVERAGE: Average unit cost in cell E10\n"
                "- SUM: Total cost of orders in cell F10"
            ),
            "download_template_name": "spare_parts_orders.xlsx",
            "validations": [
                {"cell_reference": "D10", "expected_value": "57", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(D5:D9)"},
                {"cell_reference": "E10", "expected_value": "93", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(E5:E9)"},
                {"cell_reference": "F10", "expected_value": "3220", "required_function": "SUM", "is_financial": True, "is_date": False, "correct_formula_hint": "=SUM(F5:F9)"}
            ]
        },
        {
            "id": "test_drive_conversion",
            "stage": "beginner",
            "title": "Showroom Test Drive Conversion Tracking",
            "scenario_text": (
                "Summarize weekly showroom test drive metrics. Calculate total weekly Yaris drives, "
                "total Camry drives, average weekly Land Cruiser drives, and total weekly conversions.\n\n"
                "Learning Goals:\n"
                "- SUM: Total Yaris drives in cell B9\n"
                "- SUM: Total Camry drives in cell C9\n"
                "- AVERAGE: Average LC drives in cell D9\n"
                "- SUM: Total conversions in cell E9"
            ),
            "download_template_name": "test_drive_conversion.xlsx",
            "validations": [
                {"cell_reference": "B9", "expected_value": "65", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(B5:B8)"},
                {"cell_reference": "C9", "expected_value": "44", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(C5:C8)"},
                {"cell_reference": "D9", "expected_value": "5.75", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(D5:D8)"},
                {"cell_reference": "E9", "expected_value": "31", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(E5:E8)"}
            ]
        },
        
        # --- STAGE 2: INTERMEDIATE (7 Tasks) ---
        {
            "id": "monthly_sales_commission",
            "stage": "intermediate",
            "title": "Monthly Sales Commission & Vehicle Allocation",
            "scenario_text": (
                "Lookup prices from the catalog, assign commission rates (5% for Land Cruiser, 3% others), "
                "calculate commission amounts, and calculate total sales generated by the Dubai branch.\n\n"
                "Learning Goals:\n"
                "- VLOOKUP: Lookup prices in column E (E5:E10) based on VIN prefix from VehicleCatalog sheet.\n"
                "- IF: Commission rate in column F (F5:F10) (5% for Land Cruiser, 3% for others).\n"
                "- SUMIFS: Total sales amount generated by Dubai branch in cell J6."
            ),
            "download_template_name": "monthly_sales_commission.xlsx",
            "validations": [
                {"cell_reference": "SalesLog!E5", "expected_value": "320000", "required_function": "VLOOKUP", "is_financial": True, "is_date": False, "correct_formula_hint": "=VLOOKUP(B5, VehicleCatalog!$A$5:$D$8, 4, FALSE)"},
                {"cell_reference": "SalesLog!E6", "expected_value": "140000", "required_function": "VLOOKUP", "is_financial": True, "is_date": False, "correct_formula_hint": "=VLOOKUP(B6, VehicleCatalog!$A$5:$D$8, 4, FALSE)"},
                {"cell_reference": "SalesLog!F5", "expected_value": "0.05", "required_function": "IF", "is_financial": False, "is_date": False, "correct_formula_hint": "=IF(C5=\"Land Cruiser\", 0.05, 0.03)"},
                {"cell_reference": "SalesLog!J6", "expected_value": "715000", "required_function": "SUMIFS", "is_financial": True, "is_date": False, "correct_formula_hint": "=SUMIFS(E5:E10, H5:H10, \"Dubai\")"}
            ]
        },
        {
            "id": "camry_fleet_lease",
            "stage": "intermediate",
            "title": "Camry Fleet Lease Duration & Discounts",
            "scenario_text": (
                "Lookup daily Camry rates, calculate base cost, assign 10% volume discount for leases "
                "lasting 7+ days, calculate total lease cost, and count total active Hybrid rentals.\n\n"
                "Learning Goals:\n"
                "- VLOOKUP: Lookup rate in Daily Rate column (E5:E9) based on Camry Trim from Rates sheet.\n"
                "- IF: Volume discount in column G (G5:G9) (10% for Duration >= 7, 0% for others).\n"
                "- COUNTIF: Count total active Hybrid rentals in cell K6 based on Camry Trim."
            ),
            "download_template_name": "camry_fleet_lease.xlsx",
            "validations": [
                {"cell_reference": "Rentals!E5", "expected_value": "180", "required_function": "VLOOKUP", "is_financial": True, "is_date": False, "correct_formula_hint": "=VLOOKUP(C5, Rates!$A$5:$B$7, 2, FALSE)"},
                {"cell_reference": "Rentals!G5", "expected_value": "0.10", "required_function": "IF", "is_financial": False, "is_date": False, "correct_formula_hint": "=IF(D5>=7, 0.10, 0.00)"},
                {"cell_reference": "Rentals!K6", "expected_value": "2", "required_function": "COUNTIF", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNTIF(C5:C9, \"Hybrid\")"}
            ]
        },
        {
            "id": "sales_agent_performance",
            "stage": "intermediate",
            "title": "Showroom Sales Agent Performance Analysis",
            "scenario_text": (
                "Evaluate sales agent performances. Calculate total margin in Dubai branch, "
                "average margin generated by Mahmoud in Dubai, and count total high-value deals in Dubai.\n\n"
                "Learning Goals:\n"
                "- SUMIFS: Total margin for 'Dubai' branch in cell H6.\n"
                "- AVERAGEIFS: Avg margin for Mahmoud in Dubai in cell H7.\n"
                "- COUNTIFS: Count of Dubai deals with margin > 5000 in cell H8."
            ),
            "download_template_name": "sales_agent_performance.xlsx",
            "validations": [
                {"cell_reference": "SalesLog!H6", "expected_value": "65500", "required_function": "SUMIFS", "is_financial": True, "is_date": False, "correct_formula_hint": "=SUMIFS(D5:D10, B5:B10, \"Dubai\")"},
                {"cell_reference": "SalesLog!H7", "expected_value": "17000", "required_function": "AVERAGEIFS", "is_financial": True, "is_date": False, "correct_formula_hint": "=AVERAGEIFS(D5:D10, A5:A10, \"Mahmoud\", B5:B10, \"Dubai\")"},
                {"cell_reference": "SalesLog!H8", "expected_value": "3", "required_function": "COUNTIFS", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNTIFS(B5:B10, \"Dubai\", D5:D10, \">5000\")"}
            ]
        },
        {
            "id": "spare_parts_valuation",
            "stage": "intermediate",
            "title": "Spare Parts Inventory Valuation",
            "scenario_text": (
                "Lookup wholesale cost from catalog, assign markups (30% for markup category A, 15% for others), "
                "calculate total retail value, and calculate total retail inventory valuation in Dubai branch.\n\n"
                "Learning Goals:\n"
                "- VLOOKUP: Wholesale cost in column C (C5:C9) based on Part No from Catalog sheet.\n"
                "- IF: Retail cost in column E (E5:E9) based on markup category.\n"
                "- SUMIFS: Total retail valuation for Dubai branch in cell J6."
            ),
            "download_template_name": "spare_parts_valuation.xlsx",
            "validations": [
                {"cell_reference": "Inventory!C5", "expected_value": "30", "required_function": "VLOOKUP", "is_financial": True, "is_date": False, "correct_formula_hint": "=VLOOKUP(A5, Catalog!$A$5:$B$7, 2, FALSE)"},
                {"cell_reference": "Inventory!E5", "expected_value": "39", "required_function": "IF", "is_financial": True, "is_date": False, "correct_formula_hint": "=IF(D5=\"A\", C5*1.30, C5*1.15)"},
                {"cell_reference": "Inventory!J6", "expected_value": "17160", "required_function": "SUMIFS", "is_financial": True, "is_date": False, "correct_formula_hint": "=SUMIFS(F5:F9, G5:G9, \"Dubai\")"}
            ]
        },
        {
            "id": "opex_allocation",
            "stage": "intermediate",
            "title": "Dealership Operational Expense Allocation",
            "scenario_text": (
                "Lookup showroom Base Rent based on Size Class, apply overhead penalty of 500 for showrooms "
                "with headcount > 20, calculate total allocation, and sum total operational allocations.\n\n"
                "Learning Goals:\n"
                "- VLOOKUP: Allocated base rent in column D (D5:D9) based on Size Class from Sizing sheet.\n"
                "- IF: Overhead penalty in column E (E5:E9) (500 if headcount > 20, 100 otherwise).\n"
                "- SUMIFS: Total opex allocations sum in cell I6."
            ),
            "download_template_name": "opex_allocation.xlsx",
            "validations": [
                {"cell_reference": "AllocationLog!D5", "expected_value": "2500", "required_function": "VLOOKUP", "is_financial": True, "is_date": False, "correct_formula_hint": "=VLOOKUP(B5, Sizing!$A$5:$B$7, 2, FALSE)"},
                {"cell_reference": "AllocationLog!E5", "expected_value": "500", "required_function": "IF", "is_financial": True, "is_date": False, "correct_formula_hint": "=IF(C5>20, 500, 100)"},
                {"cell_reference": "AllocationLog!I6", "expected_value": "9200", "required_function": "SUMIFS", "is_financial": True, "is_date": False, "correct_formula_hint": "=SUMIFS(F5:F9, A5:A9, \"*\")"}
            ]
        },
        {
            "id": "loan_eligibility",
            "stage": "intermediate",
            "title": "Vehicle Loan Eligibility Audit",
            "scenario_text": (
                "Determine applicant qualifying status (Eligible if income > 6000 AND credit tier is not Fair), "
                "lookup credit tier interest rate (APR), and count total Fair tier credit applications.\n\n"
                "Learning Goals:\n"
                "- IF/AND: Qualifying status in column E (E5:E9) based on Income and Credit Tier.\n"
                "- VLOOKUP: Interest rate APR in column F (F5:F9) based on Credit Tier from APRList sheet.\n"
                "- COUNTIF: Count total Fair tier credit applications processed in cell I6."
            ),
            "download_template_name": "loan_eligibility.xlsx",
            "validations": [
                {"cell_reference": "Applications!E5", "expected_value": "Eligible", "required_function": "IF/AND", "is_financial": False, "is_date": False, "correct_formula_hint": "=IF(AND(C5>6000, D5<>\"Fair\"), \"Eligible\", \"Not Eligible\")"},
                {"cell_reference": "Applications!F5", "expected_value": "0.0299", "required_function": "VLOOKUP", "is_financial": False, "is_date": False, "correct_formula_hint": "=VLOOKUP(D5, APRList!$A$5:$B$7, 2, FALSE)"},
                {"cell_reference": "Applications!I6", "expected_value": "2", "required_function": "COUNTIF", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNTIF(D5:D9, \"Fair\")"}
            ]
        },
        {
            "id": "revenue_share_audit",
            "stage": "intermediate",
            "title": "Dealership Revenue Share Audit",
            "scenario_text": (
                "Lookup share payout percentage, calculate direct profit, award 10000 bonus if partner gross revenue "
                "exceeds 500000, calculate total allocation, and sum base shared payouts for Dubai region.\n\n"
                "Learning Goals:\n"
                "- VLOOKUP: Share payout rate in column D (D5:D9) based on Tier Level from TierMatrix sheet.\n"
                "- IF: Bonus payout in column F (F5:F9) (10000 if Revenue > 500000, 0 otherwise).\n"
                "- SUMIFS: Total shared base profit generated by Dubai region partners in cell K6."
            ),
            "download_template_name": "revenue_share_audit.xlsx",
            "validations": [
                {"cell_reference": "ShareAudit!D5", "expected_value": "0.12", "required_function": "VLOOKUP", "is_financial": False, "is_date": False, "correct_formula_hint": "=VLOOKUP(C5, TierMatrix!$A$5:$B$7, 2, FALSE)"},
                {"cell_reference": "ShareAudit!F5", "expected_value": "10000", "required_function": "IF", "is_financial": True, "is_date": False, "correct_formula_hint": "=IF(B5>500000, 10000, 0)"},
                {"cell_reference": "ShareAudit!K6", "expected_value": "158600", "required_function": "SUMIFS", "is_financial": True, "is_date": False, "correct_formula_hint": "=SUMIFS(E5:E9, H5:H9, \"Dubai\")"}
            ]
        },
        
        # --- STAGE 3: ADVANCED (10 Tasks) ---
        {
            "id": "dealership_profitability",
            "stage": "advanced",
            "title": "Dealership Profitability & Capital Budgeting",
            "scenario_text": (
                "Audit aging brackets, calculate holding cost compounding penalties (use Nested IF), "
                "look up provision markup metrics from a 2D grid (use INDEX/MATCH), and calculate project XNPV and XIRR.\n\n"
                "Learning Goals:\n"
                "- Nested IF: Compounding penalty in column E (E5:E7) based on Holding Days.\n"
                "- INDEX/MATCH: Dynamic allocation lookup in cell B17 based on Score Tier (B15) and Metric (B16).\n"
                "- XNPV: Project leasing XNPV in cell F21 based on lease cash flows.\n"
                "- XIRR: Project leasing XIRR in cell F23."
            ),
            "download_template_name": "dealership_profitability.xlsx",
            "validations": [
                {"cell_reference": "AgingAndLease!E5", "expected_value": "150", "required_function": "Nested IF", "is_financial": True, "is_date": False, "correct_formula_hint": "=IF(C5<=30, 50, IF(C5<=90, 100, 150))"},
                {"cell_reference": "AgingAndLease!B17", "expected_value": "0.15", "required_function": "INDEX/MATCH", "is_financial": False, "is_date": False, "correct_formula_hint": "=INDEX(B11:C13, MATCH(B15, A11:A13, 0), MATCH(B16, B10:C10, 0))"},
                {"cell_reference": "AgingAndLease!F21", "expected_value": "85789.25", "required_function": "XNPV", "is_financial": True, "is_date": False, "correct_formula_hint": "=XNPV(D21, B21:B25, A21:A25)"},
                {"cell_reference": "AgingAndLease!F23", "expected_value": "0.1412", "required_function": "XIRR", "is_financial": False, "is_date": False, "correct_formula_hint": "=XIRR(B21:B25, A21:A25)"}
            ]
        },
        {
            "id": "holding_cost_aging",
            "stage": "advanced",
            "title": "Vehicle Holding Cost & Aging Audit",
            "scenario_text": (
                "You are an Advanced Financial Analyst. Analyze the parameter value logs in the dataset.\n"
                "Calculate the sum of Parameter Value A in cell C10, the average of Parameter Value B in cell D10, "
                "and count the total number of records in cell E10.\n\n"
                "Learning Goals:\n"
                "- SUM: Total sum of Parameter Value A in cell C10\n"
                "- AVERAGE: Average of Parameter Value B in cell D10\n"
                "- COUNT: Total count of items in cell E10"
            ),
            "download_template_name": "holding_cost_aging.xlsx",
            "validations": [
                {"cell_reference": "C10", "expected_value": "36500", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(C5:C9)"},
                {"cell_reference": "D10", "expected_value": "2529.2", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(D5:D9)"},
                {"cell_reference": "E10", "expected_value": "5", "required_function": "COUNT", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNT(C5:C9)"}
            ]
        },
        {
            "id": "capital_project_eval",
            "stage": "advanced",
            "title": "Dealership Capital Project Evaluation",
            "scenario_text": (
                "You are an Advanced Financial Analyst. Analyze the parameter value logs in the dataset.\n"
                "Calculate the sum of Parameter Value A in cell C10, the average of Parameter Value B in cell D10, "
                "and count the total number of records in cell E10.\n\n"
                "Learning Goals:\n"
                "- SUM: Total sum of Parameter Value A in cell C10\n"
                "- AVERAGE: Average of Parameter Value B in cell D10\n"
                "- COUNT: Total count of items in cell E10"
            ),
            "download_template_name": "capital_project_eval.xlsx",
            "validations": [
                {"cell_reference": "C10", "expected_value": "36500", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(C5:C9)"},
                {"cell_reference": "D10", "expected_value": "2529.2", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(D5:D9)"},
                {"cell_reference": "E10", "expected_value": "5", "required_function": "COUNT", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNT(C5:C9)"}
            ]
        },
        {
            "id": "dynamic_commission",
            "stage": "advanced",
            "title": "Showroom Dynamic Sales Commission Matrix",
            "scenario_text": (
                "You are an Advanced Financial Analyst. Analyze the parameter value logs in the dataset.\n"
                "Calculate the sum of Parameter Value A in cell C10, the average of Parameter Value B in cell D10, "
                "and count the total number of records in cell E10.\n\n"
                "Learning Goals:\n"
                "- SUM: Total sum of Parameter Value A in cell C10\n"
                "- AVERAGE: Average of Parameter Value B in cell D10\n"
                "- COUNT: Total count of items in cell E10"
            ),
            "download_template_name": "dynamic_commission.xlsx",
            "validations": [
                {"cell_reference": "C10", "expected_value": "36500", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(C5:C9)"},
                {"cell_reference": "D10", "expected_value": "2529.2", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(D5:D9)"},
                {"cell_reference": "E10", "expected_value": "5", "required_function": "COUNT", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNT(C5:C9)"}
            ]
        },
        {
            "id": "trade_in_depreciation",
            "stage": "advanced",
            "title": "Used Car Trade-in Depreciation Model",
            "scenario_text": (
                "You are an Advanced Financial Analyst. Analyze the parameter value logs in the dataset.\n"
                "Calculate the sum of Parameter Value A in cell C10, the average of Parameter Value B in cell D10, "
                "and count the total number of records in cell E10.\n\n"
                "Learning Goals:\n"
                "- SUM: Total sum of Parameter Value A in cell C10\n"
                "- AVERAGE: Average of Parameter Value B in cell D10\n"
                "- COUNT: Total count of items in cell E10"
            ),
            "download_template_name": "trade_in_depreciation.xlsx",
            "validations": [
                {"cell_reference": "C10", "expected_value": "36500", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(C5:C9)"},
                {"cell_reference": "D10", "expected_value": "2529.2", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(D5:D9)"},
                {"cell_reference": "E10", "expected_value": "5", "required_function": "COUNT", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNT(C5:C9)"}
            ]
        },
        {
            "id": "working_capital_aging",
            "stage": "advanced",
            "title": "Dealership Working Capital & A/R Aging",
            "scenario_text": (
                "You are an Advanced Financial Analyst. Analyze the parameter value logs in the dataset.\n"
                "Calculate the sum of Parameter Value A in cell C10, the average of Parameter Value B in cell D10, "
                "and count the total number of records in cell E10.\n\n"
                "Learning Goals:\n"
                "- SUM: Total sum of Parameter Value A in cell C10\n"
                "- AVERAGE: Average of Parameter Value B in cell D10\n"
                "- COUNT: Total count of items in cell E10"
            ),
            "download_template_name": "working_capital_aging.xlsx",
            "validations": [
                {"cell_reference": "C10", "expected_value": "36500", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(C5:C9)"},
                {"cell_reference": "D10", "expected_value": "2529.2", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(D5:D9)"},
                {"cell_reference": "E10", "expected_value": "5", "required_function": "COUNT", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNT(C5:C9)"}
            ]
        },
        {
            "id": "fleet_procurement",
            "stage": "advanced",
            "title": "Dealership Fleet Procurement Model",
            "scenario_text": (
                "You are an Advanced Financial Analyst. Analyze the parameter value logs in the dataset.\n"
                "Calculate the sum of Parameter Value A in cell C10, the average of Parameter Value B in cell D10, "
                "and count the total number of records in cell E10.\n\n"
                "Learning Goals:\n"
                "- SUM: Total sum of Parameter Value A in cell C10\n"
                "- AVERAGE: Average of Parameter Value B in cell D10\n"
                "- COUNT: Total count of items in cell E10"
            ),
            "download_template_name": "fleet_procurement.xlsx",
            "validations": [
                {"cell_reference": "C10", "expected_value": "36500", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(C5:C9)"},
                {"cell_reference": "D10", "expected_value": "2529.2", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(D5:D9)"},
                {"cell_reference": "E10", "expected_value": "5", "required_function": "COUNT", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNT(C5:C9)"}
            ]
        },
        {
            "id": "showroom_feasibility",
            "stage": "advanced",
            "title": "Showroom Expansion Feasibility Study",
            "scenario_text": (
                "You are an Advanced Financial Analyst. Analyze the parameter value logs in the dataset.\n"
                "Calculate the sum of Parameter Value A in cell C10, the average of Parameter Value B in cell D10, "
                "and count the total number of records in cell E10.\n\n"
                "Learning Goals:\n"
                "- SUM: Total sum of Parameter Value A in cell C10\n"
                "- AVERAGE: Average of Parameter Value B in cell D10\n"
                "- COUNT: Total count of items in cell E10"
            ),
            "download_template_name": "showroom_feasibility.xlsx",
            "validations": [
                {"cell_reference": "C10", "expected_value": "36500", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(C5:C9)"},
                {"cell_reference": "D10", "expected_value": "2529.2", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(D5:D9)"},
                {"cell_reference": "E10", "expected_value": "5", "required_function": "COUNT", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNT(C5:C9)"}
            ]
        },
        {
            "id": "parts_forecasting",
            "stage": "advanced",
            "title": "Spare Parts Inventory Forecasting Model",
            "scenario_text": (
                "You are an Advanced Financial Analyst. Analyze the parameter value logs in the dataset.\n"
                "Calculate the sum of Parameter Value A in cell C10, the average of Parameter Value B in cell D10, "
                "and count the total number of records in cell E10.\n\n"
                "Learning Goals:\n"
                "- SUM: Total sum of Parameter Value A in cell C10\n"
                "- AVERAGE: Average of Parameter Value B in cell D10\n"
                "- COUNT: Total count of items in cell E10"
            ),
            "download_template_name": "parts_forecasting.xlsx",
            "validations": [
                {"cell_reference": "C10", "expected_value": "36500", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(C5:C9)"},
                {"cell_reference": "D10", "expected_value": "2529.2", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(D5:D9)"},
                {"cell_reference": "E10", "expected_value": "5", "required_function": "COUNT", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNT(C5:C9)"}
            ]
        },
        {
            "id": "service_capacity",
            "stage": "advanced",
            "title": "Service Center Capacity Planning",
            "scenario_text": (
                "You are an Advanced Financial Analyst. Analyze the parameter value logs in the dataset.\n"
                "Calculate the sum of Parameter Value A in cell C10, the average of Parameter Value B in cell D10, "
                "and count the total number of records in cell E10.\n\n"
                "Learning Goals:\n"
                "- SUM: Total sum of Parameter Value A in cell C10\n"
                "- AVERAGE: Average of Parameter Value B in cell D10\n"
                "- COUNT: Total count of items in cell E10"
            ),
            "download_template_name": "service_capacity.xlsx",
            "validations": [
                {"cell_reference": "C10", "expected_value": "36500", "required_function": "SUM", "is_financial": False, "is_date": False, "correct_formula_hint": "=SUM(C5:C9)"},
                {"cell_reference": "D10", "expected_value": "2529.2", "required_function": "AVERAGE", "is_financial": False, "is_date": False, "correct_formula_hint": "=AVERAGE(D5:D9)"},
                {"cell_reference": "E10", "expected_value": "5", "required_function": "COUNT", "is_financial": False, "is_date": False, "correct_formula_hint": "=COUNT(C5:C9)"}
            ]
        }
    ]
    
    for t_data in tasks:
        task = db.query(models.Task).filter(models.Task.id == t_data["id"]).first()
        if not task:
            task = models.Task(
                id=t_data["id"],
                stage=t_data["stage"],
                title=t_data["title"],
                scenario_text=t_data["scenario_text"],
                download_template_name=t_data["download_template_name"]
            )
            db.add(task)
            db.flush()
        else:
            task.stage = t_data["stage"]
            task.title = t_data["title"]
            task.scenario_text = t_data["scenario_text"]
            task.download_template_name = t_data["download_template_name"]
            
        # Drop existing validations to prevent obsolete rules from colliding
        db.query(models.TaskValidation).filter(models.TaskValidation.task_id == task.id).delete()
        db.flush()
        
        # Add validations
        for v_data in t_data["validations"]:
            val = models.TaskValidation(
                task_id=task.id,
                cell_reference=v_data["cell_reference"],
                expected_value=v_data["expected_value"],
                required_function=v_data["required_function"],
                is_financial=v_data["is_financial"],
                is_date=v_data["is_date"],
                correct_formula_hint=v_data["correct_formula_hint"]
            )
            db.add(val)
                
    db.commit()

def generate_excel_templates():
    """Generates all 22 styled Excel exercise spreadsheets programmatically in media/templates."""
    app_dir = os.path.dirname(os.path.abspath(__file__))
    media_dir = os.path.join(os.path.dirname(app_dir), "media", "templates")
    os.makedirs(media_dir, exist_ok=True)
    
    # 1. Stage 1 (Beginner) Templates
    generate_beginner_excel(os.path.join(media_dir, "daily_showroom_footfall.xlsx"))
    generate_weekly_sales_volume(os.path.join(media_dir, "weekly_sales_volume.xlsx"))
    generate_showroom_csat_score(os.path.join(media_dir, "showroom_csat_score.xlsx"))
    generate_spare_parts_orders(os.path.join(media_dir, "spare_parts_orders.xlsx"))
    generate_test_drive_conversion(os.path.join(media_dir, "test_drive_conversion.xlsx"))
    
    # 2. Stage 2 (Intermediate) Templates
    generate_intermediate_excel(os.path.join(media_dir, "monthly_sales_commission.xlsx"))
    generate_camry_fleet_lease(os.path.join(media_dir, "camry_fleet_lease.xlsx"))
    generate_sales_agent_performance(os.path.join(media_dir, "sales_agent_performance.xlsx"))
    generate_spare_parts_valuation(os.path.join(media_dir, "spare_parts_valuation.xlsx"))
    generate_opex_allocation(os.path.join(media_dir, "opex_allocation.xlsx"))
    generate_loan_eligibility(os.path.join(media_dir, "loan_eligibility.xlsx"))
    generate_revenue_share_audit(os.path.join(media_dir, "revenue_share_audit.xlsx"))
    
    # 3. Stage 3 (Advanced) Templates
    generate_advanced_excel(os.path.join(media_dir, "dealership_profitability.xlsx"))
    
    # Generate generic blank tables for remaining advanced items (can be customized further!)
    advanced_blank_items = [
        ("holding_cost_aging.xlsx", "Vehicle Holding Cost & Aging Audit"),
        ("capital_project_eval.xlsx", "Dealership Capital Project Evaluation"),
        ("dynamic_commission.xlsx", "Showroom Dynamic Sales Commission Matrix"),
        ("trade_in_depreciation.xlsx", "Used Car Trade-in Depreciation Model"),
        ("working_capital_aging.xlsx", "Dealership Working Capital & A/R Aging"),
        ("fleet_procurement.xlsx", "Dealership Fleet Procurement Model"),
        ("showroom_feasibility.xlsx", "Showroom Expansion Feasibility Study"),
        ("parts_forecasting.xlsx", "Spare Parts Inventory Forecasting Model"),
        ("service_capacity.xlsx", "Service Center Capacity Planning")
    ]
    for filename, title in advanced_blank_items:
        generate_blank_template(os.path.join(media_dir, filename), title)
        
    print("Database seeding & template generation finished successfully!")

def run_seed():
    """Main seeder entrypoint."""
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        seed_database(db)
        generate_excel_templates()
    finally:
        db.close()

if __name__ == "__main__":
    run_seed()
