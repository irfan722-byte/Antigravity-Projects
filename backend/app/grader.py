import re
import datetime
import io
import openpyxl
from typing import Dict, List, Tuple, Any

def parse_excel_date(cell_val: Any) -> str:
    """Helper to convert cell values to standard DD.MM.YYYY string."""
    if isinstance(cell_val, (datetime.datetime, datetime.date)):
        return cell_val.strftime("%d.%m.%Y")
    
    if isinstance(cell_val, str):
        # Clean string
        cleaned = cell_val.strip()
        # Check if already in DD.MM.YYYY
        if re.match(r"^\d{2}\.\d{2}\.\d{4}$", cleaned):
            return cleaned
        
        # Try to parse standard ISO format YYYY-MM-DD
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                dt = datetime.datetime.strptime(cleaned, fmt)
                return dt.strftime("%d.%m.%Y")
            except ValueError:
                continue
                
    return str(cell_val).strip()

def parse_numeric_value(val: Any) -> float:
    """Helper to parse any value into a float, stripping common currency signs/spaces."""
    if isinstance(val, (int, float)):
        return float(val)
    if val is None:
        return 0.0
    
    # Strip spaces, commas, currency signs like AED, $, etc.
    cleaned = str(val).upper().replace(",", "").replace("AED", "").replace("$", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0

def check_function_usage(formula: str, required_func: str) -> Tuple[bool, str]:
    """
    Verify if the required function exists inside the Excel formula.
    Supports complex combinations like INDEX/MATCH or Nested IFs.
    """
    formula_upper = formula.upper()
    req_upper = required_func.upper().strip()
    
    if not req_upper:
        return True, ""
        
    # Support 'INDEX/MATCH' or 'INDEX, MATCH' - both must be present
    if "/" in req_upper or "," in req_upper:
        funcs = [f.strip() for f in re.split(r"[/,]", req_upper)]
        for f in funcs:
            if f not in formula_upper:
                return False, f"Missing required function '{f}'"
        return True, ""
        
    # Support 'NESTED IF' or 'NESTED IFS' - check for at least 2 occurrences of 'IF'
    if req_upper in ("NESTED IF", "NESTED IFS"):
        # Count the occurrences of 'IF(' in the formula
        if formula_upper.count("IF(") < 2:
            return False, "Formula must use Nested IF statements (minimum 2 levels)"
        return True, ""
        
    # Standard single function check (e.g. 'SUM', 'VLOOKUP', 'XNPV')
    # Use word boundary or simple in check with parenthesis/delimiters to avoid matching within words
    # e.g., we want to match 'IF' but not 'LIFETIME'
    pattern = rf"\b{re.escape(req_upper)}\b|\b{re.escape(req_upper)}\("
    if not re.search(pattern, formula_upper) and req_upper not in formula_upper:
        return False, f"Missing required function '{required_func}'"
        
    return True, ""

def grade_excel_file(file_contents: bytes, validations: List[Any]) -> Tuple[bool, Dict[str, Any]]:
    """
    Grades the uploaded Excel workbook against a set of validation rules.
    Reads values via data_only=True and formulas via data_only=False.
    """
    try:
        # Load values workbook
        wb_val = openpyxl.load_workbook(io.BytesIO(file_contents), data_only=True)
        # Load formulas workbook
        wb_form = openpyxl.load_workbook(io.BytesIO(file_contents), data_only=False)
    except Exception as e:
        return False, {
            "error": f"Invalid Excel workbook file structure: {str(e)}"
        }
        
    errors = {}
    passed_count = 0
    total_validations = len(validations)
    
    for val in validations:
        ref = val.cell_reference.strip()
        # Parse sheet and coordinate, e.g. "SalesLog!E4" or "E4" (default to active sheet)
        if "!" in ref:
            sheet_name, cell_coord = ref.split("!", 1)
        else:
            sheet_name = None
            cell_coord = ref
            
        try:
            # Resolve sheets
            if sheet_name:
                if sheet_name not in wb_val.sheetnames:
                    errors[ref] = {
                        "expected_value": val.expected_value,
                        "required_function": val.required_function or "None",
                        "correct_formula_hint": val.correct_formula_hint or "",
                        "actual_value": "N/A",
                        "actual_formula": "N/A",
                        "error_message": f"Sheet '{sheet_name}' not found in workbook"
                    }
                    continue
                sheet_val = wb_val[sheet_name]
                sheet_form = wb_form[sheet_name]
            else:
                sheet_val = wb_val.active
                sheet_form = wb_form.active
                
            cell_val = sheet_val[cell_coord]
            cell_form = sheet_form[cell_coord]
            
            actual_value = cell_val.value
            actual_formula = str(cell_form.value or "")
            
            # 1. Formula Check: must start with '='
            if not actual_formula.startswith("="):
                errors[ref] = {
                    "expected_value": val.expected_value,
                    "required_function": val.required_function or "None",
                    "correct_formula_hint": val.correct_formula_hint or "",
                    "actual_value": str(actual_value) if actual_value is not None else "Empty",
                    "actual_formula": actual_formula if actual_formula else "Value (Hardcoded)",
                    "error_message": "Hardcoded value detected. You must enter an Excel formula starting with '='"
                }
                continue
                
            # 2. Function Check: must use the required Excel function
            if val.required_function:
                has_func, err_msg = check_function_usage(actual_formula, val.required_function)
                if not has_func:
                    errors[ref] = {
                        "expected_value": val.expected_value,
                        "required_function": val.required_function,
                        "correct_formula_hint": val.correct_formula_hint or "",
                        "actual_value": str(actual_value) if actual_value is not None else "Empty",
                        "actual_formula": actual_formula,
                        "error_message": err_msg
                    }
                    continue
                    
            # 3. Value Match Validation
            value_matches = False
            error_message = ""
            
            if val.is_financial:
                # Standardize to 2 decimal places
                try:
                    act_num = parse_numeric_value(actual_value)
                    exp_num = parse_numeric_value(val.expected_value)
                    if abs(act_num - exp_num) < 0.015:  # Tolerate small float/rounding discrepancies
                        value_matches = True
                    else:
                        value_matches = False
                        error_message = f"Value mismatch. Expected AED {exp_num:.2f}, got AED {act_num:.2f}"
                except Exception:
                    value_matches = False
                    error_message = f"Value mismatch. Expected financial value {val.expected_value}"
                    
            elif val.is_date:
                # Format to DD.MM.YYYY string
                try:
                    act_date = parse_excel_date(actual_value)
                    exp_date = parse_excel_date(val.expected_value)
                    if act_date == exp_date:
                        value_matches = True
                    else:
                        value_matches = False
                        error_message = f"Date mismatch. Expected {exp_date}, got {act_date}"
                except Exception:
                    value_matches = False
                    error_message = f"Date mismatch. Expected date {val.expected_value}"
            else:
                # Normal comparison
                if actual_value is not None:
                    # Let's try matching float/int if possible
                    try:
                        act_num = float(str(actual_value).strip())
                        exp_num = float(str(val.expected_value).strip())
                        if abs(act_num - exp_num) < 0.015:
                            value_matches = True
                        else:
                            value_matches = False
                            error_message = f"Value mismatch. Expected {exp_num}, got {act_num}"
                    except ValueError:
                        # Fallback to string comparison
                        act_str = str(actual_value).strip().lower()
                        exp_str = str(val.expected_value).strip().lower()
                        if act_str == exp_str:
                            value_matches = True
                        else:
                            value_matches = False
                            error_message = f"Value mismatch. Expected '{val.expected_value}', got '{actual_value}'"
                else:
                    value_matches = False
                    error_message = "Cell is empty"
                    
            if not value_matches:
                errors[ref] = {
                    "expected_value": val.expected_value,
                    "required_function": val.required_function or "None",
                    "correct_formula_hint": val.correct_formula_hint or "",
                    "actual_value": str(actual_value) if actual_value is not None else "Empty",
                    "actual_formula": actual_formula,
                    "error_message": error_message
                }
            else:
                passed_count += 1
                
        except Exception as cell_err:
            errors[ref] = {
                "expected_value": val.expected_value,
                "required_function": val.required_function or "None",
                "correct_formula_hint": val.correct_formula_hint or "",
                "actual_value": "Error",
                "actual_formula": "Error",
                "error_message": f"Fatal error checking cell: {str(cell_err)}"
            }
            
    is_success = (passed_count == total_validations)
    return is_success, errors
